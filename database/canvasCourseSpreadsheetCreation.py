import sys
sys.path.append("../")
from canvaslib.Canvaslib import databaseSearchCanvas, parseArgsCanvas, initialiseLogging
from canvaslib.CanvasAPIClass import CanvasAPI
try:
    from simple_settings import settings
except:
    from lib.CanvasSettings import CanvasSettings
    settings = CanvasSettings()
import logging
from pprint import pprint
from openpyxl import Workbook, worksheet

def mainProgram():
    
    args, myCanvas = setupVariables()
    
    print('Collecting Course Data From Canvas')
    courseDict = populateCourseSite(myCanvas)
    
    fileName = f'{settings.CANVAS_FOLDER}\\Canvas Course List.xlsx'
    outputToSpreadsheet(courseDict, fileName)
    
    print(f'Course List written to {fileName}')

    
def outputToSpreadsheet(courseDict, fileName):
    
    workbook = Workbook()
    
    sheet = workbook.active
    
    count = 1
    
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 60
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 25
    sheet.column_dimensions['H'].width = 10
    
    for val, text in enumerate(['Canvas Course ID', 'Account ID', 'Account Name', 'Course Name', 'Course sis Id', 'Course Code', 'Course Term', 'Course Term ID']):
        sheet.cell(count, (val+1)).value = text
        
    for val, (key, value) in enumerate(courseDict.items()):
        count += 1
        for val, text in enumerate([key, value['accountId'], value['accountName'], value['courseName'], value['courseSisId'], value['courseCode'], value['courseTerm'], value['courseTermId']]):
            sheet.cell(count, (val+1)).value = text
            
    mediumStyle = worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    # create a table
    table = worksheet.table.Table(ref=f'A1:H{len(courseDict)+1}', tableStyleInfo=mediumStyle, displayName='CanvasCourseList')
    # add the table to the worksheet
    sheet.add_table(table)
    
    workbook.save(fileName)

def populateCourseSite(myCanvas):
    
    entries = {}
    accounts = myCanvas.getAccounts()
    
    include = ['term']
    
    for account in accounts:
        courses = account.get_courses(include=include)
        
        for course in courses:
            entries[course.id] = {'courseName' : course.name,
                                  'courseSisId' : course.sis_course_id,
                                  'courseCode' : course.course_code,
                                  'courseTerm' : course.term['name'],
                                  'courseTermId' : course.enrollment_term_id,
                                  'accountId' : account.id,
                                  'accountName' : account.name,
                                  }
    return entries


def setupVariables():
        
    initialiseLogging("course spreadsheet generation.log")
    
    args = parseArgsCanvas()
    
    return args, CanvasAPI()  
    
if __name__ == '__main__':
    
    mainProgram()


