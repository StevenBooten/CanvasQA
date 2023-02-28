import sys
import progressbar
sys.path.append("../")
try:
    from simple_settings import settings
except:
    from lib.CanvasSettings import CanvasSettings
    settings = CanvasSettings()
from canvaslib.Canvaslib import databaseSearchCanvas, parseArgsCanvas, initialiseLogging
from canvaslib.CanvasAPIClass import CanvasAPI
from Checks.body import checkPageBody
from Gatherers.assignments import collectCourseAssignments
from Gatherers.fileStructure import collectCourseFiles
from Gatherers.moduleInfo import collectCourseModules, unattachedPages, collectCoursePages
from htmlgeneration.canvasQaHtml import generateQaHtml
import json
from pprint import pprint
from pathlib import Path
from time import sleep
import time
from openpyxl import load_workbook


sys.setrecursionlimit(2000)

MAXBAR = 100
def updateBar(count, bar):
    step = MAXBAR / 8 #This is the number of steps in the main program
    
    count += 1
    
    bar.update((MAXBAR/step) * count)

def mainProgram():
     
    args, courseDetails, myCanvas = setupVariables()
    
    
    # This input takes a single course code or a comma separated list of course codes.
    # can either be course code ie 1270 or the full url ie https://lms.griffith.edu.au/courses/1270 seperated by commas (,)
    if args.url is not None:
        urlList = args.url.split(',')
        
        courseDetails = []
        for url in urlList:
            if url.startswith('https://'):
                url = url.replace(f'{settings.CANVAS_API_URL}courses/', '')
                url = url[:url.find('/')]
                
            if not url.isdigit():
                print(f'Invalid course code: {url}')
                continue
            
            courseDetails.append({'canvasCourseId': url})
            
    if args.spreadsheet is not None:
        courseDetails = pullDataFromSpreadsheet(args.spreadsheet)
            
    
    for prog, course in enumerate(courseDetails):
        sys.setrecursionlimit(10000)
        count = 1
        canvasQa = {}
        canvasQa['usedFiles'] = []
        canvasQa['issues'] = {}
        canvasQa['issues']['Images'] = { 'id':"collapsible-img-check", 'count' : 0, 'html' : ''}
        canvasQa['issues']['Assignments'] = { 'id':"collapsible-assignment-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['Blackboard Residuals'] = { 'id':"collapsible-bb-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['File Structure'] = {'id': "collapsible-filestructure-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['Modules'] = { 'id':"collapsible-modules-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['Placeholders'] = { 'id':"collapsible-placeholder-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['Unattached Pages'] = { 'id':"collapsible-unattachedpages-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['Course Links'] = { 'id':"collapsible-link-check", 'count' : 0, 'html' : '' }
        canvasQa['issues']['Embedded Content'] = { 'id':"collapsible-video-check", 'count' : 0, 'html' : ''}
        
        myCanvas.getCourse(course['canvasCourseId'])
        print('-' * 80)
        print(f'Running: {myCanvas.courseCode}')
        print(f'{prog+1} of {len(courseDetails)}')
        sys.stdout.flush()
        bar = progressbar.ProgressBar(maxval=MAXBAR, \
            widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
        
        bar.start()
        
        canvasQa['pages'] = collectCoursePages(myCanvas)
        updateBar(count, bar)
        
        canvasQa['usedFiles'] += collectCourseModules(myCanvas, canvasQa)
        updateBar(count, bar)
        
        usedFiles = collectCourseAssignments(myCanvas, canvasQa)
        canvasQa['usedFiles'] += usedFiles
        updateBar(count, bar)
        
        canvasQa['unattachedPages'] = unattachedPages(myCanvas, canvasQa)
        updateBar(count, bar)
        
        usedFiles = checkPageBody(canvasQa, myCanvas)
        canvasQa['usedFiles'] += usedFiles
        updateBar(count, bar)
        
        canvasQa['files'] = collectCourseFiles(myCanvas, canvasQa)
        updateBar(count, bar)
        
        filename = f'{myCanvas.courseCode.replace("_"," ").replace("/", "-").replace(":", "-").replace("?", "").replace("*.*", "")}'
        
        canvasQaHtml = generateQaHtml(myCanvas, canvasQa)
        saveQaHtml(canvasQaHtml, myCanvas, filename)
        #updateBar(count, bar)
        
        
        
        with open(f'{settings.CANVAS_QA_DOWNLOAD_FOLDER}\\jsons\\{filename} QA.json', 'w') as outfile:
            json.dump(canvasQa, outfile, indent=4)
        updateBar(count, bar)
            
        
        
        bar.finish()
        sys.stdout.flush()
        
    if args.loop:
        return 1
    else:
        return 0

def pullDataFromSpreadsheet(spreadsheet):
    courseDetails = []
    
    workbook = load_workbook(filename=spreadsheet)
    
    sheet = workbook.active
    
    for row in sheet.iter_rows():
        if row[0].value == 'Canvas Course ID':
            continue
        courseDetails.append({'canvasCourseId' : row[0].value}) 
    
    return courseDetails
    
def saveQaHtml(canvasQaHtml, myCanvas, filename, attempt=0):
    
    filePath = f'{settings.CANVAS_QA_DOWNLOAD_FOLDER}'
    fileName =  f'{filename} QA v2.html'
    
    # uploads the HTML report to the canvas course site
    with open(Path(filePath, fileName), 'w', encoding="utf-8") as f:
        f.write(canvasQaHtml.render(pretty=True, doctype=True))
    try:    
        myCanvas.uploadFile(filePath, fileName, '/QA Info')
    except:
        if attempt < 1:
            saveQaHtml(canvasQaHtml, myCanvas, filename, 1)
        else:
            pass

def setupVariables():
        
    initialiseLogging("Canvas QA.log")
    
    args = parseArgsCanvas()   
    
    courseList = ['7917QCA']
        
    courseDetails = databaseSearchCanvas(args.group, args.accountId, courseList=courseList, oua=args.oua, list=args.list, term=args.term, year=args.year, course=args.course, ignoreJoined=args.ignoreJoined)
    
    return args, courseDetails, CanvasAPI()
       
        
    
if __name__ == '__main__':
    runProgram = 1
    
    while runProgram:
        runProgram = mainProgram() 
    